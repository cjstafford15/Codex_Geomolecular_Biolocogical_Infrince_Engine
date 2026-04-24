from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "data" / "geomolecular_template.xlsx"


SHEETS: dict[str, list[str]] = {
    "BiologicalDefinitions": [
        "bio_id",
        "term",
        "category",
        "definition",
        "confirmation_role",
        "safety_notes",
        "source_ids",
    ],
    "GeometrySignatures": [
        "geometry_id",
        "geometry_name",
        "section",
        "shape_class",
        "symmetry",
        "topology",
        "scale",
        "polarity_pattern",
        "dynamics",
        "match_features",
        "exclusion_features",
    ],
    "CandidateMaterials": [
        "candidate_id",
        "name",
        "candidate_type",
        "geometry_ids",
        "bio_ids",
        "notes",
        "safety_status",
        "source_ids",
    ],
    "Observations": [
        "observation_id",
        "candidate_id",
        "geometry_id",
        "bio_id",
        "context",
        "observed_pattern",
        "confidence",
        "contradiction",
        "source_ids",
    ],
    "Sources": [
        "source_id",
        "title",
        "source_type",
        "url_or_path",
        "reliability",
        "notes",
    ],
    "ProductIdeas": [
        "idea_id",
        "title",
        "geometry_ids",
        "candidate_ids",
        "intended_domain",
        "concept_note",
        "required_validation",
        "blocked_claims",
    ],
}


SEED_ROWS: dict[str, list[list[str]]] = {
    "BiologicalDefinitions": [
        [
            "BIO-FLY-ATTRACTION",
            "Fly attraction",
            "behavior",
            "Movement or aggregation of flies toward a stimulus.",
            "Confirms whether a geometry-linked volatile, surface, or temporal pattern has behavioral relevance.",
            "Do not infer product efficacy without controlled observation and safety review.",
            "SRC-PROJECT-NOTE-001",
        ],
        [
            "BIO-GNAT-FLIGHT-REDUCTION",
            "Indoor gnat flight reduction",
            "behavior",
            "Observed reduction in flying gnat presence in a defined indoor context.",
            "Confirms whether a geometry-linked environmental or volatile pattern changes observed movement.",
            "Do not create pesticide directions or harmful application instructions.",
            "SRC-PROJECT-NOTE-001",
        ],
        [
            "BIO-SURFACE-ADHESION-REPEL",
            "Surface adhesion repulsion",
            "behavior",
            "Organism avoids or cannot maintain contact with a treated surface.",
            "Confirms whether a geometry-linked surface film or interface pattern changes contact behavior.",
            "Do not infer pesticidal or medical surface treatment efficacy.",
            "SRC-PROJECT-NOTE-001",
        ],
        [
            "BIO-MOSQUITO-ATTRACTION",
            "Mosquito attraction",
            "behavior",
            "Movement of mosquitoes toward a carbon dioxide or skin-emission mimic.",
            "Confirms whether a geometry-linked breath-mimic or skin-odor plume attracts mosquitoes.",
            "Do not infer repellency inverse without separate validation.",
            "SRC-HOUSEHOLD-002",
        ],
        [
            "BIO-RECEPTOR-POCKET-FIT",
            "Receptor pocket geometric fit",
            "pharmacology-proxy",
            "Hypothetical binding based on complementary shape and polarity between molecule and receptor cavity.",
            "Confirms whether a geometry-linked molecular shape could theoretically fit a known receptor pocket.",
            "This is geometric hypothesis only; no pharmacological claim or dosing is implied.",
            "SRC-MEDICAL-003",
        ],
        [
            "BIO-SKIN-BARRIER-PASSAGE",
            "Skin barrier passage geometry",
            "dermatology-proxy",
            "Observed or hypothesized permeation through stratum corneum lipid matrix.",
            "Confirms whether a geometry-linked amphiphilic or small-polar pattern could traverse skin layers.",
            "No medical delivery claim without permeation measurement and safety review.",
            "SRC-MEDICAL-003",
        ],
    ],
    "GeometrySignatures": [
        [
            "GEO-VOLATILE-PLUME-GRADIENT",
            "Volatile plume gradient",
            "Volatile Plumes",
            "gradient",
            "asymmetric",
            "branching diffusion field",
            "room to organism",
            "volatile polarity gradient",
            "evaporative diffusion",
            "Odor intensity changes over distance; airflow and source surface alter the pattern.",
            "No measurable volatile gradient or behavior occurs only by visual/contact cues.",
        ],
        [
            "GEO-VOLATILE-PLUME-PULSE",
            "Volatile plume temporal pulse",
            "Volatile Plumes",
            "waveform",
            "periodic",
            "pulsed release nodes",
            "room to organism",
            "polarity oscillation",
            "intermittent release",
            "Odor releases in bursts rather than steady stream; temporal pattern may trigger different behavior than constant exposure.",
            "Behavior identical between steady and pulsed release at same average concentration.",
        ],
        [
            "GEO-SURFACE-FILM-INTERFACE",
            "Surface film interface",
            "Surface Films",
            "plane",
            "asymmetric",
            "two-phase interface",
            "surface to organism",
            "amphiphilic boundary",
            "adhesive/static",
            "Thin film, meniscus, residue, or coating changes contact behavior at an interface.",
            "Effect occurs without contact or surface interaction.",
        ],
        [
            "GEO-BINDING-POCKET-CAVITY",
            "Binding pocket cavity",
            "Binding Pockets",
            "cavity",
            "chiral",
            "enclosed pocket with channel",
            "molecular",
            "hydrophobic pocket with polar rim",
            "static/lock-and-key",
            "Complementary shape and polarity distribution allow molecular docking into a receptor cavity.",
            "No measurable shape complementarity or binding affinity data available.",
        ],
        [
            "GEO-SKIN-LAMELLAR-LAYER",
            "Skin lamellar lipid layer",
            "Surface Films",
            "layered lattice",
            "bilayer periodic",
            "stacked sheets with defects",
            "cellular",
            "amphiphilic lamellae",
            "static/diffusive",
            "Lipid bilayers arranged in stacked sheets; hydrophilic and hydrophobic regions alternate. Permeation depends on defect geometry and molecule aspect ratio.",
            "No lipid bilayer structure or skin barrier model available.",
        ],
        [
            "GEO-TEMPORAL-PULSE-TRAIN",
            "Temporal pulse train",
            "Temporal Pulses",
            "waveform",
            "periodic",
            "discrete time series",
            "time-series",
            "on/off polarity",
            "oscillating",
            "Repetitive stimulus bursts at fixed or variable intervals; entrainment or habituation may depend on frequency and duty cycle.",
            "Continuous stimulus produces identical behavior.",
        ],
        [
            "GEO-SURFACE-TEXTURE-RIDGE",
            "Surface texture ridge",
            "Surface Films",
            "line array",
            "periodic",
            "parallel ridges with grooves",
            "surface to organism",
            "hydrophobic ridge / hydrophilic groove",
            "static/topographic",
            "Micro-ridges or grooves change wettability and contact angle; affects droplet retention or organism attachment.",
            "Smooth surface produces identical contact behavior.",
        ],
    ],
    "CandidateMaterials": [
        [
            "CAND-APPLE-CIDER-VINEGAR",
            "Apple cider vinegar",
            "food",
            "GEO-VOLATILE-PLUME-GRADIENT",
            "BIO-FLY-ATTRACTION",
            "Keep as food-derived candidate data, not as a finalized product instruction.",
            "needs review",
            "SRC-PROJECT-NOTE-001",
        ],
        [
            "CAND-SUGAR-FERMENT-AROMA",
            "Sugar fermentation aroma",
            "recipe",
            "GEO-VOLATILE-PLUME-GRADIENT",
            "BIO-FLY-ATTRACTION",
            "Recipe-like records are allowed when stored as observations and geometry signatures.",
            "needs review",
            "SRC-PROJECT-NOTE-001",
        ],
        [
            "CAND-YEAST-BLOOM-MASH",
            "Yeast bloom mash",
            "recipe",
            "GEO-VOLATILE-PLUME-GRADIENT,GEO-VOLATILE-PLUME-PULSE",
            "BIO-FLY-ATTRACTION,BIO-MOSQUITO-ATTRACTION",
            "Fermented grain mash with active yeast produces intermittent CO2 and volatile bursts. Geometry combines steady plume and temporal pulse.",
            "needs review",
            "SRC-RECIPE-004",
        ],
        [
            "CAND-ESSENTIAL-OIL-DIFFUSER",
            "Essential oil diffuser",
            "household",
            "GEO-VOLATILE-PLUME-PULSE",
            "BIO-GNAT-FLIGHT-REDUCTION",
            "Ultrasonic diffuser releases essential oils in intermittent bursts; temporal pulse geometry may differ from steady evaporative source.",
            "needs review",
            "SRC-HOUSEHOLD-002",
        ],
        [
            "CAND-SOAP-FILM-RESIDUE",
            "Soap film residue",
            "household",
            "GEO-SURFACE-FILM-INTERFACE",
            "BIO-SURFACE-ADHESION-REPEL",
            "Thin soap film left on surface changes contact angle and may reduce insect landing or adhesion.",
            "needs review",
            "SRC-HOUSEHOLD-002",
        ],
        [
            "CAND-CITRONELLA-CANDLE",
            "Citronella candle",
            "household",
            "GEO-VOLATILE-PLUME-GRADIENT,GEO-TEMPORAL-PULSE-TRAIN",
            "BIO-MOSQUITO-ATTRACTION,BIO-GNAT-FLIGHT-REDUCTION",
            "Burning candle produces steady volatile gradient and periodic heat-driven convection pulses. Geometry is compound.",
            "needs review",
            "SRC-HOUSEHOLD-002",
        ],
        [
            "CAND-CAPSAICIN-EXTRACT",
            "Capsaicin extract",
            "botanical",
            "GEO-BINDING-POCKET-CAVITY",
            "BIO-RECEPTOR-POCKET-FIT",
            "Capsaicin has a vanillyl group and hydrophobic tail that may fit TRPV1 receptor pocket geometry.",
            "needs review",
            "SRC-MEDICAL-003",
        ],
        [
            "CAND-MENTHOL-CRYSTAL",
            "Menthol crystal",
            "botanical",
            "GEO-BINDING-POCKET-CAVITY,GEO-SKIN-LAMELLAR-LAYER",
            "BIO-RECEPTOR-POCKET-FIT,BIO-SKIN-BARRIER-PASSAGE",
            "Menthol is small, amphiphilic, and has a cyclohexanol ring that may fit TRPM8 geometry while traversing skin lipid lamellae.",
            "needs review",
            "SRC-MEDICAL-003",
        ],
        [
            "CAND-ALOE-VERA-GEL",
            "Aloe vera gel",
            "botanical",
            "GEO-SKIN-LAMELLAR-LAYER",
            "BIO-SKIN-BARRIER-PASSAGE",
            "Polysaccharide gel may create a surface film that alters skin barrier geometry and moisture retention.",
            "needs review",
            "SRC-MEDICAL-003",
        ],
        [
            "CAND-BAKING-SODA-PASTE",
            "Baking soda paste",
            "household",
            "GEO-SURFACE-TEXTURE-RIDGE",
            "BIO-SURFACE-ADHESION-REPEL",
            "Powder-in-water paste dries to a micro-ridged residue that may alter surface contact geometry.",
            "low concern",
            "SRC-HOUSEHOLD-002",
        ],
    ],
    "Observations": [
        [
            "OBS-001",
            "CAND-APPLE-CIDER-VINEGAR",
            "GEO-VOLATILE-PLUME-GRADIENT",
            "BIO-FLY-ATTRACTION",
            "Household anecdote placeholder",
            "Attraction appears linked to evaporating food odor geometry.",
            "low",
            "No controlled comparison yet.",
            "SRC-PROJECT-NOTE-001",
        ],
        [
            "OBS-002",
            "CAND-YEAST-BLOOM-MASH",
            "GEO-VOLATILE-PLUME-PULSE",
            "BIO-MOSQUITO-ATTRACTION",
            "Backyard fermentation experiment",
            "Mosquitoes appeared in waves correlated with visible CO2 bubble bursts from the mash.",
            "low",
            "No trap or counterfactual; may be attracted to heat, moisture, or other cues.",
            "SRC-RECIPE-004",
        ],
        [
            "OBS-003",
            "CAND-ESSENTIAL-OIL-DIFFUSER",
            "GEO-VOLATILE-PLUME-PULSE",
            "BIO-GNAT-FLIGHT-REDUCTION",
            "Kitchen counter observation",
            "Gnat presence near fruit bowl reduced when diffuser was active on intermittent mode; steady mode had weaker apparent effect.",
            "low",
            "No blinded or replicated trial; humidity change confounder.",
            "SRC-HOUSEHOLD-002",
        ],
        [
            "OBS-004",
            "CAND-SOAP-FILM-RESIDUE",
            "GEO-SURFACE-FILM-INTERFACE",
            "BIO-SURFACE-ADHESION-REPEL",
            "Bathroom mirror cleaning test",
            "Flies appeared to avoid landing on mirror surface with visible soap streaks compared to untreated mirror.",
            "low",
            "Mirror reflection confounder; no controlled surface pair.",
            "SRC-HOUSEHOLD-002",
        ],
        [
            "OBS-005",
            "CAND-MENTHOL-CRYSTAL",
            "GEO-BINDING-POCKET-CAVITY",
            "BIO-RECEPTOR-POCKET-FIT",
            "Literature-derived hypothesis",
            "Menthol cyclohexanol ring size and hydroxyl position match known TRPM8 pocket geometry in docking models.",
            "medium",
            "In silico only; no wet-lab binding assay.",
            "SRC-MEDICAL-003",
        ],
        [
            "OBS-006",
            "CAND-MENTHOL-CRYSTAL",
            "GEO-SKIN-LAMELLAR-LAYER",
            "BIO-SKIN-BARRIER-PASSAGE",
            "Transdermal patch literature review",
            "Menthol enhances permeation of co-applied compounds through stratum corneum; may create defects in lipid lamellae.",
            "medium",
            "Mechanism not fully resolved; may also act via keratinocyte disruption.",
            "SRC-MEDICAL-003",
        ],
        [
            "OBS-007",
            "CAND-CITRONELLA-CANDLE",
            "GEO-TEMPORAL-PULSE-TRAIN",
            "BIO-GNAT-FLIGHT-REDUCTION",
            "Patio evening observation",
            "Gnat annoyance seemed lower near lit citronella candle than near unlit candle of same size and shape.",
            "low",
            "Heat, light, and convection confounders; no chemical-only control.",
            "SRC-HOUSEHOLD-002",
        ],
    ],
    "Sources": [
        [
            "SRC-PROJECT-NOTE-001",
            "Initial project framing",
            "project note",
            "local",
            "unverified",
            "Seed source used to demonstrate schema only.",
        ],
        [
            "SRC-HOUSEHOLD-002",
            "Household observation collection",
            "observation",
            "local",
            "anecdotal",
            "Anecdotal household notes collected during informal observation periods.",
        ],
        [
            "SRC-MEDICAL-003",
            "Medical pharmacology hypothesis notes",
            "hypothesis",
            "local",
            "unverified",
            "Hypothesis notes for geometric docking and dermal passage; not clinical evidence.",
        ],
        [
            "SRC-RECIPE-004",
            "Fermentation recipe observation log",
            "recipe",
            "local",
            "anecdotal",
            "Observations made during home fermentation experiments; not controlled.",
        ],
    ],
    "ProductIdeas": [
        [
            "IDEA-GEOMETRY-LED-INSECT-BEHAVIOR",
            "Geometry-led insect behavior screen",
            "GEO-VOLATILE-PLUME-GRADIENT,GEO-SURFACE-FILM-INTERFACE",
            "CAND-APPLE-CIDER-VINEGAR,CAND-SUGAR-FERMENT-AROMA",
            "household research",
            "Compare candidate materials by geometry signature before biological category.",
            "Controlled observation, exposure safety review, environmental review, and claim review.",
            "Do not claim attraction, repellency, elimination, medical value, or pesticide efficacy yet.",
        ],
        [
            "IDEA-PULSE-MODULATED-AROMA",
            "Pulse modulated aroma delivery",
            "GEO-VOLATILE-PLUME-PULSE,GEO-TEMPORAL-PULSE-TRAIN",
            "CAND-ESSENTIAL-OIL-DIFFUSER,CAND-YEAST-BLOOM-MASH,CAND-CITRONELLA-CANDLE",
            "household research",
            "Investigate whether intermittent volatile release produces different behavioral outcomes than steady release for the same total emitted mass.",
            "Controlled observation with matched total volatile output; temporal pattern as independent variable.",
            "Do not claim repellency, medical effect, or guaranteed behavioral change.",
        ],
        [
            "IDEA-GEOMETRY-LED-SURFACE-COATING",
            "Geometry-led surface coating screen",
            "GEO-SURFACE-FILM-INTERFACE,GEO-SURFACE-TEXTURE-RIDGE",
            "CAND-SOAP-FILM-RESIDUE,CAND-BAKING-SODA-PASTE",
            "material science",
            "Compare surface residues and dried films by contact-angle geometry and micro-topography before testing biological adhesion.",
            "Surface profilometry, contact angle measurement, biological adhesion assay, safety review.",
            "Do not claim pesticidal, medical, or antimicrobial surface action without validation.",
        ],
        [
            "IDEA-DERMAL-GEOMETRY-SCREEN",
            "Dermal permeation geometry screen",
            "GEO-SKIN-LAMELLAR-LAYER,GEO-BINDING-POCKET-CAVITY",
            "CAND-MENTHOL-CRYSTAL,CAND-ALOE-VERA-GEL,CAND-CAPSAICIN-EXTRACT",
            "medical research",
            "Compare botanical candidates by skin-barrier passage geometry and receptor-pocket fit geometry before any pharmacological interpretation.",
            "Franz cell permeation assay, receptor binding assay, safety and irritation review.",
            "Do not claim therapeutic effect, dosing, or clinical efficacy.",
        ],
    ],
}


def add_table(ws, sheet_name: str, row_count: int, column_count: int) -> None:
    end_column = ws.cell(row=1, column=column_count).column_letter
    table = Table(displayName=f"{sheet_name}Table", ref=f"A1:{end_column}{row_count}")
    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table.tableStyleInfo = style
    ws.add_table(table)


def build_workbook() -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)

    for sheet_name, headers in SHEETS.items():
        ws = wb.create_sheet(sheet_name)
        ws.append(headers)
        for row in SEED_ROWS.get(sheet_name, []):
            ws.append(row)

        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")

        for column_cells in ws.columns:
            header = str(column_cells[0].value)
            width = max(14, min(42, len(header) + 6))
            ws.column_dimensions[column_cells[0].column_letter].width = width

        ws.freeze_panes = "A2"
        add_table(ws, sheet_name, ws.max_row, len(headers))

    return wb


def main() -> None:
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    workbook = build_workbook()
    workbook.save(OUTPUT)
    print(f"Created {OUTPUT}")


if __name__ == "__main__":
    main()
