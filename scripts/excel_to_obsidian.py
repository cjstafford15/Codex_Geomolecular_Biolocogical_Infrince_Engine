from __future__ import annotations

import argparse
import re
from collections import defaultdict
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


REQUIRED_SHEETS = {
    "BiologicalDefinitions",
    "GeometrySignatures",
    "CandidateMaterials",
    "Observations",
    "Sources",
    "ProductIdeas",
}


def clean(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def split_ids(value: str) -> list[str]:
    return [part.strip() for part in value.split(",") if part.strip()]


def slugify(value: str) -> str:
    text = value.strip().lower()
    text = re.sub(r"[^a-z0-9]+", "-", text)
    return text.strip("-") or "untitled"


def note_link(title: str) -> str:
    return f"[[{title}]]"


def bullet_links(titles: Iterable[str]) -> str:
    values = [title for title in titles if title]
    if not values:
        return "- None recorded"
    return "\n".join(f"- {note_link(title)}" for title in sorted(set(values)))


def read_sheet(workbook_path: Path, sheet_name: str) -> list[dict[str, str]]:
    wb = load_workbook(workbook_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        return []

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [clean(header) for header in rows[0]]
    records: list[dict[str, str]] = []
    for row in rows[1:]:
        record = {headers[index]: clean(value) for index, value in enumerate(row) if index < len(headers)}
        if any(record.values()):
            records.append(record)
    return records


def load_records(workbook_path: Path) -> dict[str, list[dict[str, str]]]:
    wb = load_workbook(workbook_path, read_only=True)
    missing = REQUIRED_SHEETS.difference(wb.sheetnames)
    if missing:
        missing_list = ", ".join(sorted(missing))
        raise ValueError(f"Workbook is missing required sheet(s): {missing_list}")
    wb.close()
    return {sheet_name: read_sheet(workbook_path, sheet_name) for sheet_name in REQUIRED_SHEETS}


def write_note(path: Path, content: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(content.rstrip() + "\n", encoding="utf-8")


def index_by(records: list[dict[str, str]], key: str) -> dict[str, dict[str, str]]:
    return {record[key]: record for record in records if record.get(key)}


def render_geometry_notes(records: dict[str, list[dict[str, str]]], vault: Path) -> None:
    geometries = records["GeometrySignatures"]
    candidates = records["CandidateMaterials"]
    observations = records["Observations"]
    bios = records["BiologicalDefinitions"]

    bio_by_id = index_by(bios, "bio_id")

    candidates_by_geometry: dict[str, list[dict[str, str]]] = defaultdict(list)
    bio_ids_by_geometry: dict[str, set[str]] = defaultdict(set)
    for candidate in candidates:
        for geometry_id in split_ids(candidate.get("geometry_ids", "")):
            candidates_by_geometry[geometry_id].append(candidate)
            bio_ids_by_geometry[geometry_id].update(split_ids(candidate.get("bio_ids", "")))

    observations_by_geometry: dict[str, list[dict[str, str]]] = defaultdict(list)
    for observation in observations:
        geometry_id = observation.get("geometry_id", "")
        if geometry_id:
            observations_by_geometry[geometry_id].append(observation)
            if observation.get("bio_id"):
                bio_ids_by_geometry[geometry_id].add(observation["bio_id"])

    for geometry in geometries:
        geometry_id = geometry["geometry_id"]
        title = geometry.get("geometry_name") or geometry_id
        section = geometry.get("section") or "Unsorted Geometry"
        candidate_titles = [candidate.get("name", "") for candidate in candidates_by_geometry[geometry_id]]
        observation_titles = [
            f"{observation.get('observation_id', 'Observation')} - {observation.get('observed_pattern', '')[:70]}"
            for observation in observations_by_geometry[geometry_id]
        ]
        bio_titles = [
            bio_by_id[bio_id].get("term", bio_id)
            for bio_id in bio_ids_by_geometry[geometry_id]
            if bio_id in bio_by_id
        ]

        content = f"""# {title}

Type: `geometry-section`
Geometry ID: `{geometry_id}`
Section: `{section}`

## Geometry

- Shape class: {geometry.get("shape_class", "")}
- Symmetry: {geometry.get("symmetry", "")}
- Topology: {geometry.get("topology", "")}
- Scale: {geometry.get("scale", "")}
- Polarity pattern: {geometry.get("polarity_pattern", "")}
- Dynamics: {geometry.get("dynamics", "")}

## Match Features

{geometry.get("match_features", "") or "None recorded"}

## Exclusion Features

{geometry.get("exclusion_features", "") or "None recorded"}

## Candidate Materials

{bullet_links(candidate_titles)}

## Observations

{bullet_links(observation_titles)}

## Biological Confirmation Terms

{bullet_links(bio_titles)}
"""
        write_note(vault / "Geometry" / section / f"{slugify(title)}.md", content)


def render_candidate_notes(records: dict[str, list[dict[str, str]]], vault: Path) -> None:
    geometries = index_by(records["GeometrySignatures"], "geometry_id")
    bios = index_by(records["BiologicalDefinitions"], "bio_id")
    sources = index_by(records["Sources"], "source_id")

    for candidate in records["CandidateMaterials"]:
        title = candidate.get("name") or candidate["candidate_id"]
        geometry_titles = [
            geometries[geometry_id].get("geometry_name", geometry_id)
            for geometry_id in split_ids(candidate.get("geometry_ids", ""))
            if geometry_id in geometries
        ]
        bio_titles = [
            bios[bio_id].get("term", bio_id)
            for bio_id in split_ids(candidate.get("bio_ids", ""))
            if bio_id in bios
        ]
        source_titles = [
            sources[source_id].get("title", source_id)
            for source_id in split_ids(candidate.get("source_ids", ""))
            if source_id in sources
        ]

        content = f"""# {title}

Type: `candidate-material`
Candidate ID: `{candidate.get("candidate_id", "")}`
Candidate type: `{candidate.get("candidate_type", "")}`
Safety status: `{candidate.get("safety_status", "")}`

## Geometry Links

{bullet_links(geometry_titles)}

## Biological Confirmation Links

{bullet_links(bio_titles)}

## Notes

{candidate.get("notes", "") or "None recorded"}

## Sources

{bullet_links(source_titles)}
"""
        write_note(vault / "Candidates" / f"{slugify(title)}.md", content)


def render_biology_notes(records: dict[str, list[dict[str, str]]], vault: Path) -> None:
    sources = index_by(records["Sources"], "source_id")

    for bio in records["BiologicalDefinitions"]:
        title = bio.get("term") or bio["bio_id"]
        source_titles = [
            sources[source_id].get("title", source_id)
            for source_id in split_ids(bio.get("source_ids", ""))
            if source_id in sources
        ]
        content = f"""# {title}

Type: `biological-confirmation`
Biology ID: `{bio.get("bio_id", "")}`
Category: `{bio.get("category", "")}`

## Definition

{bio.get("definition", "") or "None recorded"}

## Confirmation Role

{bio.get("confirmation_role", "") or "None recorded"}

## Safety Notes

{bio.get("safety_notes", "") or "None recorded"}

## Sources

{bullet_links(source_titles)}
"""
        write_note(vault / "Biology Confirmation" / f"{slugify(title)}.md", content)


def render_observation_notes(records: dict[str, list[dict[str, str]]], vault: Path) -> None:
    candidates = index_by(records["CandidateMaterials"], "candidate_id")
    geometries = index_by(records["GeometrySignatures"], "geometry_id")
    bios = index_by(records["BiologicalDefinitions"], "bio_id")

    for observation in records["Observations"]:
        observation_id = observation.get("observation_id", "Observation")
        title = f"{observation_id} - {observation.get('observed_pattern', '')[:70]}".strip(" -")
        candidate = candidates.get(observation.get("candidate_id", ""), {})
        geometry = geometries.get(observation.get("geometry_id", ""), {})
        bio = bios.get(observation.get("bio_id", ""), {})
        content = f"""# {title}

Type: `observation`
Observation ID: `{observation_id}`
Confidence: `{observation.get("confidence", "")}`

## Links

- Candidate: {note_link(candidate.get("name", observation.get("candidate_id", "")))}
- Geometry: {note_link(geometry.get("geometry_name", observation.get("geometry_id", "")))}
- Biological confirmation: {note_link(bio.get("term", observation.get("bio_id", "")))}

## Context

{observation.get("context", "") or "None recorded"}

## Observed Pattern

{observation.get("observed_pattern", "") or "None recorded"}

## Contradiction

{observation.get("contradiction", "") or "None recorded"}
"""
        write_note(vault / "Observations" / f"{slugify(title)}.md", content)


def render_product_notes(records: dict[str, list[dict[str, str]]], vault: Path) -> None:
    geometries = index_by(records["GeometrySignatures"], "geometry_id")
    candidates = index_by(records["CandidateMaterials"], "candidate_id")

    for idea in records["ProductIdeas"]:
        title = idea.get("title") or idea["idea_id"]
        geometry_titles = [
            geometries[geometry_id].get("geometry_name", geometry_id)
            for geometry_id in split_ids(idea.get("geometry_ids", ""))
            if geometry_id in geometries
        ]
        candidate_titles = [
            candidates[candidate_id].get("name", candidate_id)
            for candidate_id in split_ids(idea.get("candidate_ids", ""))
            if candidate_id in candidates
        ]
        content = f"""# {title}

Type: `product-concept`
Idea ID: `{idea.get("idea_id", "")}`
Intended domain: `{idea.get("intended_domain", "")}`

## Geometry Drivers

{bullet_links(geometry_titles)}

## Candidate Inputs

{bullet_links(candidate_titles)}

## Concept Note

{idea.get("concept_note", "") or "None recorded"}

## Required Validation

{idea.get("required_validation", "") or "None recorded"}

## Blocked Claims

{idea.get("blocked_claims", "") or "None recorded"}
"""
        write_note(vault / "Product Ideas" / f"{slugify(title)}.md", content)


def render_index(vault: Path) -> None:
    content = """# Geomolecular Biological Inference Engine

Geometry is the primary matching layer. Biology is the confirmation layer.

## Sections

- [[Geometry]]
- [[Candidates]]
- [[Biology Confirmation]]
- [[Observations]]
- [[Product Ideas]]

## Agent Focus

- Search by geometry signatures before biological terms.
- Treat recipes, foods, and household observations as candidate records.
- Use biological evidence to confirm, reject, or weaken a geometry match.
- Keep product and medical ideas non-actionable until validation and safety review exist.
"""
    write_note(vault / "Home.md", content)


def export(workbook: Path, vault: Path) -> None:
    records = load_records(workbook)
    render_index(vault)
    render_geometry_notes(records, vault)
    render_candidate_notes(records, vault)
    render_biology_notes(records, vault)
    render_observation_notes(records, vault)
    render_product_notes(records, vault)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Export geomolecular Excel workbook to an Obsidian vault.")
    parser.add_argument("--workbook", type=Path, required=True, help="Path to the Excel workbook.")
    parser.add_argument("--vault", type=Path, required=True, help="Output directory for Obsidian markdown files.")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    export(args.workbook.resolve(), args.vault.resolve())
    print(f"Exported {args.workbook} to {args.vault}")


if __name__ == "__main__":
    main()

