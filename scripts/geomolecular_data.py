from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


CORE_SHEETS = {
    "BiologicalDefinitions",
    "GeometrySignatures",
    "CandidateMaterials",
    "Observations",
    "Sources",
    "ProductIdeas",
}

OPTIONAL_SHEETS = {
    "ExternalAgents",
    "ShadowAgents",
}

ID_COLUMNS = {
    "BiologicalDefinitions": "bio_id",
    "GeometrySignatures": "geometry_id",
    "CandidateMaterials": "candidate_id",
    "Observations": "observation_id",
    "Sources": "source_id",
    "ProductIdeas": "idea_id",
    "ExternalAgents": "agent_id",
    "ShadowAgents": "shadow_agent_id",
}

CONFIDENCE_SCORES = {
    "high": 1.0,
    "medium": 0.6,
    "low": 0.3,
}

RELIABILITY_SCORES = {
    "peer-reviewed": 1.0,
    "high": 0.9,
    "medium": 0.65,
    "dataset": 0.65,
    "simulation-framework": 0.55,
    "paper": 0.8,
    "observation": 0.4,
    "anecdotal": 0.35,
    "recipe": 0.3,
    "project note": 0.35,
    "hypothesis": 0.4,
    "low": 0.25,
    "unverified": 0.25,
    "unknown": 0.0,
    "": 0.0,
}


@dataclass(frozen=True)
class ValidationIssue:
    severity: str
    sheet: str
    record_id: str
    message: str


def clean(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def split_ids(value: str) -> list[str]:
    return [part.strip() for part in value.split(",") if part.strip()]


def index_by(records: list[dict[str, str]], key: str) -> dict[str, dict[str, str]]:
    return {record[key]: record for record in records if record.get(key)}


def read_workbook(
    workbook_path: Path,
    required_sheets: Iterable[str] = CORE_SHEETS,
    optional_sheets: Iterable[str] = OPTIONAL_SHEETS,
) -> dict[str, list[dict[str, str]]]:
    wb = load_workbook(workbook_path, data_only=True, read_only=True)
    try:
        required = set(required_sheets)
        optional = set(optional_sheets)
        missing = required.difference(wb.sheetnames)
        if missing:
            missing_list = ", ".join(sorted(missing))
            raise ValueError(f"Workbook is missing required sheet(s): {missing_list}")

        records: dict[str, list[dict[str, str]]] = {}
        for sheet_name in sorted(required.union(optional)):
            if sheet_name in wb.sheetnames:
                records[sheet_name] = _read_sheet_from_workbook(wb, sheet_name)
            else:
                records[sheet_name] = []
        return records
    finally:
        wb.close()


def _read_sheet_from_workbook(wb, sheet_name: str) -> list[dict[str, str]]:
    ws = wb[sheet_name]
    rows = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows)
    except StopIteration:
        return []

    headers = [clean(header) for header in header_row]
    records: list[dict[str, str]] = []
    for row in rows:
        record = {
            headers[index]: clean(value)
            for index, value in enumerate(row)
            if index < len(headers) and headers[index]
        }
        if any(record.values()):
            records.append(record)
    return records


def confidence_score(value: str) -> float:
    normalized = clean(value).lower()
    if normalized in CONFIDENCE_SCORES:
        return CONFIDENCE_SCORES[normalized]
    try:
        numeric = float(normalized)
    except ValueError:
        return 0.0
    if numeric > 1:
        numeric = numeric / 100
    return max(0.0, min(1.0, numeric))


def reliability_score(source: dict[str, str] | None) -> float:
    if not source:
        return 0.0
    reliability = clean(source.get("reliability", "")).lower()
    source_type = clean(source.get("source_type", "")).lower()
    return max(
        RELIABILITY_SCORES.get(reliability, 0.0),
        RELIABILITY_SCORES.get(source_type, 0.0),
    )


def average(values: Iterable[float]) -> float:
    materialized = list(values)
    if not materialized:
        return 0.0
    return round(sum(materialized) / len(materialized), 3)


def average_source_reliability(
    source_ids: str,
    sources_by_id: dict[str, dict[str, str]],
) -> float:
    return average(reliability_score(sources_by_id.get(source_id)) for source_id in split_ids(source_ids))


def has_meaningful_contradiction(value: str) -> bool:
    text = clean(value).lower()
    if not text:
        return False
    return text not in {"none", "none recorded", "n/a", "na", "no contradiction"}


def validation_score(confidence: float, source_reliability: float, contradiction: bool) -> float:
    # A recorded confounder reduces readiness, but it is better than an untracked gap.
    penalty = 0.1 if contradiction else 0.0
    score = (confidence * 0.55) + (source_reliability * 0.45) - penalty
    return round(max(0.0, min(1.0, score)), 3)


def validate_records(records: dict[str, list[dict[str, str]]]) -> list[ValidationIssue]:
    issues: list[ValidationIssue] = []
    indexes = {
        sheet: index_by(records.get(sheet, []), column)
        for sheet, column in ID_COLUMNS.items()
    }

    for sheet, column in ID_COLUMNS.items():
        seen: set[str] = set()
        for record in records.get(sheet, []):
            record_id = record.get(column, "")
            if not record_id:
                issues.append(ValidationIssue("error", sheet, "", f"Missing required ID column {column}."))
            elif record_id in seen:
                issues.append(ValidationIssue("error", sheet, record_id, f"Duplicate {column}."))
            seen.add(record_id)

    _check_source_refs(issues, records, indexes)
    _check_candidate_refs(issues, records, indexes)
    _check_observation_refs(issues, records, indexes)
    _check_product_refs(issues, records, indexes)
    _check_external_agent_refs(issues, records, indexes)
    _check_confidence_values(issues, records)

    return issues


def _check_source_refs(
    issues: list[ValidationIssue],
    records: dict[str, list[dict[str, str]]],
    indexes: dict[str, dict[str, dict[str, str]]],
) -> None:
    source_ids = set(indexes["Sources"])
    for sheet in ("BiologicalDefinitions", "CandidateMaterials", "Observations", "ProductIdeas", "ExternalAgents"):
        for record in records.get(sheet, []):
            record_id = _record_id(sheet, record)
            for source_id in split_ids(record.get("source_ids", "")):
                if source_id not in source_ids:
                    issues.append(ValidationIssue("error", sheet, record_id, f"Unknown source_id {source_id}."))


def _check_candidate_refs(
    issues: list[ValidationIssue],
    records: dict[str, list[dict[str, str]]],
    indexes: dict[str, dict[str, dict[str, str]]],
) -> None:
    geometry_ids = set(indexes["GeometrySignatures"])
    bio_ids = set(indexes["BiologicalDefinitions"])
    for record in records.get("CandidateMaterials", []):
        record_id = record.get("candidate_id", "")
        for geometry_id in split_ids(record.get("geometry_ids", "")):
            if geometry_id not in geometry_ids:
                issues.append(ValidationIssue("error", "CandidateMaterials", record_id, f"Unknown geometry_id {geometry_id}."))
        for bio_id in split_ids(record.get("bio_ids", "")):
            if bio_id not in bio_ids:
                issues.append(ValidationIssue("error", "CandidateMaterials", record_id, f"Unknown bio_id {bio_id}."))


def _check_observation_refs(
    issues: list[ValidationIssue],
    records: dict[str, list[dict[str, str]]],
    indexes: dict[str, dict[str, dict[str, str]]],
) -> None:
    candidate_ids = set(indexes["CandidateMaterials"])
    geometry_ids = set(indexes["GeometrySignatures"])
    bio_ids = set(indexes["BiologicalDefinitions"])
    for record in records.get("Observations", []):
        record_id = record.get("observation_id", "")
        if record.get("candidate_id", "") not in candidate_ids:
            issues.append(ValidationIssue("error", "Observations", record_id, f"Unknown candidate_id {record.get('candidate_id', '')}."))
        if record.get("geometry_id", "") not in geometry_ids:
            issues.append(ValidationIssue("error", "Observations", record_id, f"Unknown geometry_id {record.get('geometry_id', '')}."))
        if record.get("bio_id", "") not in bio_ids:
            issues.append(ValidationIssue("error", "Observations", record_id, f"Unknown bio_id {record.get('bio_id', '')}."))


def _check_product_refs(
    issues: list[ValidationIssue],
    records: dict[str, list[dict[str, str]]],
    indexes: dict[str, dict[str, dict[str, str]]],
) -> None:
    candidate_ids = set(indexes["CandidateMaterials"])
    geometry_ids = set(indexes["GeometrySignatures"])
    for record in records.get("ProductIdeas", []):
        record_id = record.get("idea_id", "")
        for geometry_id in split_ids(record.get("geometry_ids", "")):
            if geometry_id not in geometry_ids:
                issues.append(ValidationIssue("error", "ProductIdeas", record_id, f"Unknown geometry_id {geometry_id}."))
        for candidate_id in split_ids(record.get("candidate_ids", "")):
            if candidate_id not in candidate_ids:
                issues.append(ValidationIssue("error", "ProductIdeas", record_id, f"Unknown candidate_id {candidate_id}."))


def _check_external_agent_refs(
    issues: list[ValidationIssue],
    records: dict[str, list[dict[str, str]]],
    indexes: dict[str, dict[str, dict[str, str]]],
) -> None:
    agent_ids = set(indexes["ExternalAgents"])
    shadow_ids = set(indexes["ShadowAgents"])
    candidate_ids = set(indexes["CandidateMaterials"])

    for record in records.get("ExternalAgents", []):
        record_id = record.get("agent_id", "")
        shadow_id = record.get("shadow_agent_id", "")
        if shadow_id and shadow_id not in shadow_ids:
            issues.append(ValidationIssue("error", "ExternalAgents", record_id, f"Unknown shadow_agent_id {shadow_id}."))
        for candidate_id in split_ids(record.get("candidate_links", "")):
            if candidate_id not in candidate_ids:
                issues.append(ValidationIssue("warning", "ExternalAgents", record_id, f"Unknown candidate_link {candidate_id}."))

    for record in records.get("ShadowAgents", []):
        record_id = record.get("shadow_agent_id", "")
        primary_agent_id = record.get("primary_agent_id", "")
        if primary_agent_id and primary_agent_id not in agent_ids:
            issues.append(ValidationIssue("error", "ShadowAgents", record_id, f"Unknown primary_agent_id {primary_agent_id}."))


def _check_confidence_values(
    issues: list[ValidationIssue],
    records: dict[str, list[dict[str, str]]],
) -> None:
    for record in records.get("Observations", []):
        value = record.get("confidence", "")
        if value and confidence_score(value) == 0.0:
            issues.append(ValidationIssue("warning", "Observations", record.get("observation_id", ""), f"Unrecognized confidence value {value}."))


def _record_id(sheet: str, record: dict[str, str]) -> str:
    id_column = ID_COLUMNS.get(sheet, "")
    return record.get(id_column, "")
