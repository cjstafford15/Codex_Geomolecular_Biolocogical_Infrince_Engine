from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "data" / "geomolecular_template.xlsx"
VAULT = ROOT / "obsidian-vault"
REPORT_DIR = VAULT / "Agent Reports"


def clean(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def split_ids(value: str) -> list[str]:
    return [part.strip() for part in value.split(",") if part.strip()]


def read_sheet(workbook_path: Path, sheet_name: str) -> list[dict[str, str]]:
    wb = load_workbook(workbook_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [clean(h) for h in rows[0]]
    records = []
    for row in rows[1:]:
        record = {headers[i]: clean(v) for i, v in enumerate(row) if i < len(headers)}
        if any(record.values()):
            records.append(record)
    return records


def load_records(workbook_path: Path) -> dict[str, list[dict[str, str]]]:
    sheets = {
        "GeometrySignatures",
        "CandidateMaterials",
        "Observations",
        "BiologicalDefinitions",
        "ProductIdeas",
        "Sources",
    }
    return {s: read_sheet(workbook_path, s) for s in sheets}


def generate_concepts(records: dict[str, list[dict[str, str]]]) -> list[dict]:
    geometries = records["GeometrySignatures"]
    candidates = records["CandidateMaterials"]
    observations = records["Observations"]
    existing_ideas = records["ProductIdeas"]

    geo_by_id = {g["geometry_id"]: g for g in geometries if g.get("geometry_id")}
    cand_by_id = {c["candidate_id"]: c for c in candidates if c.get("candidate_id")}

    concepts = []

    # Strategy 1: Candidates sharing the same geometry
    geo_to_candidates: dict[str, list[str]] = {}
    for c in candidates:
        for gid in split_ids(c.get("geometry_ids", "")):
            geo_to_candidates.setdefault(gid, []).append(c["candidate_id"])

    for gid, cand_ids in geo_to_candidates.items():
        if len(cand_ids) < 2:
            continue
        geo = geo_by_id.get(gid)
        if not geo:
            continue

        names = [cand_by_id[cid].get("name", cid) for cid in cand_ids if cid in cand_by_id]

        concepts.append({
            "concept_id": f"CONCEPT-COMBINE-{gid}",
            "strategy": "combine_same_geometry",
            "geometry_id": gid,
            "geometry_name": geo.get("geometry_name", gid),
            "candidate_names": names,
            "concept_note": (
                f"Candidates {names} all express geometry signature "
                f"'{geo.get('geometry_name', gid)}'. "
                f"Consider combining or comparing features across these candidates."
            ),
        })

    # Strategy 2: Geometries with observations but few candidates
    obs_by_geo: dict[str, int] = {}
    for o in observations:
        gid = o.get("geometry_id", "")
        if gid:
            obs_by_geo[gid] = obs_by_geo.get(gid, 0) + 1

    for gid, count in obs_by_geo.items():
        cands = geo_to_candidates.get(gid, [])
        if len(cands) < 2 and count >= 1:
            geo = geo_by_id.get(gid)
            if not geo:
                continue
            concepts.append({
                "concept_id": f"CONCEPT-GAP-{gid}",
                "strategy": "fill_geometry_gap",
                "geometry_id": gid,
                "geometry_name": geo.get("geometry_name", gid),
                "candidate_names": [cand_by_id[c].get("name", c) for c in cands],
                "concept_note": (
                    f"Geometry '{geo.get('geometry_name', gid)}' has {count} observation(s) "
                    f"but only {len(cands)} candidate(s). Consider searching for more candidates."
                ),
            })

    # Strategy 3: Extend existing product ideas
    for idea in existing_ideas:
        idea_id = idea.get("idea_id", "")
        title = idea.get("title", "")
        geom_ids = split_ids(idea.get("geometry_ids", ""))
        cand_ids = split_ids(idea.get("candidate_ids", ""))

        geo_names = [geo_by_id[g].get("geometry_name", g) for g in geom_ids if g in geo_by_id]
        cand_names = [cand_by_id[c].get("name", c) for c in cand_ids if c in cand_by_id]

        if geo_names and cand_names:
            concepts.append({
                "concept_id": f"CONCEPT-IDEA-{idea_id}",
                "strategy": "extend_existing_idea",
                "geometry_id": ",".join(geom_ids),
                "geometry_name": ", ".join(geo_names),
                "candidate_names": cand_names,
                "concept_note": (
                    f"Existing idea '{title}' is driven by geometries: {', '.join(geo_names)}. "
                    f"Consider additional candidates that share these geometries."
                ),
            })

    return concepts


def render_report(concepts: list[dict]) -> str:
    lines = [
        "# Product Concept Report",
        "",
        "Type: `agent-report`",
        "Agent: `product-concept-agent`",
        "",
        "## Summary",
        "",
        f"- Concepts generated: {len(concepts)}",
        "",
        "## Concepts",
        "",
    ]

    for c in concepts:
        lines.append(f"### {c['concept_id']}")
        lines.append("")
        lines.append(f"- Strategy: `{c['strategy']}`")
        lines.append(f"- Geometry: [[{c['geometry_name']}]] (`{c['geometry_id']}`)")
        lines.append("")
        lines.append("**Candidates:**")
        for name in c["candidate_names"]:
            lines.append(f"- [[{name}]]")
        lines.append("")
        lines.append("**Concept Note:**")
        lines.append(c["concept_note"])
        lines.append("")
        lines.append("---")
        lines.append("")

    if not concepts:
        lines.append("No concepts generated. Add more candidates, geometries, or observations.")
        lines.append("")

    return "\n".join(lines)


def main() -> None:
    records = load_records(WORKBOOK)
    concepts = generate_concepts(records)

    REPORT_DIR.mkdir(parents=True, exist_ok=True)

    json_path = REPORT_DIR / "product-concept-report.json"
    json_path.write_text(json.dumps(concepts, indent=2), encoding="utf-8")

    md_path = REPORT_DIR / "product-concept-report.md"
    md_path.write_text(render_report(concepts), encoding="utf-8")

    print(f"Report written to {md_path}")


if __name__ == "__main__":
    main()
