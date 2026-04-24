from __future__ import annotations

import json
from collections import defaultdict
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "data" / "geomolecular_template.xlsx"
VAULT = ROOT / "obsidian-vault"
REPORT_DIR = VAULT / "Agent Reports"


def clean(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def split_ids(value: str) -> list[str]:
    return [part.strip() for part in value.split(",") if part.strip()]


def read_sheet(workbook_path: Path, sheet_name: str) -> list[dict[str, str]]:
    wb = load_workbook(workbook_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [clean(h) for h in rows[0]]
    records = []
    for row in rows[1:]:
        record = {headers[i]: clean(v) for i, v in enumerate(row) if i < len(headers)}
        if any(record.values()):
            records.append(record)
    return records


def load_records(workbook_path: Path) -> dict[str, list[dict[str, str]]]:
    sheets = {
        "GeometrySignatures",
        "CandidateMaterials",
        "Observations",
        "BiologicalDefinitions",
        "ProductIdeas",
    }
    return {s: read_sheet(workbook_path, s) for s in sheets}


def analyze_patterns(records: dict[str, list[dict[str, str]]]) -> dict:
    geometries = records["GeometrySignatures"]
    candidates = records["CandidateMaterials"]
    observations = records["Observations"]
    bios = records["BiologicalDefinitions"]

    bio_by_id = {b["bio_id"]: b for b in bios if b.get("bio_id")}
    geo_by_id = {g["geometry_id"]: g for g in geometries if g.get("geometry_id")}

    # Group candidates by geometry
    candidates_by_geo: dict[str, list[dict]] = defaultdict(list)
    for c in candidates:
        for gid in split_ids(c.get("geometry_ids", "")):
            candidates_by_geo[gid].append(c)

    # Group observations by geometry
    observations_by_geo: dict[str, list[dict]] = defaultdict(list)
    for o in observations:
        gid = o.get("geometry_id", "")
        if gid:
            observations_by_geo[gid].append(o)

    # Find cross-biology geometry matches (same geometry, different biology)
    cross_bio_matches = []
    for gid, geo in geo_by_id.items():
        cand_list = candidates_by_geo.get(gid, [])
        obs_list = observations_by_geo.get(gid, [])

        bio_ids = set()
        for c in cand_list:
            bio_ids.update(split_ids(c.get("bio_ids", "")))
        for o in obs_list:
            if o.get("bio_id"):
                bio_ids.add(o["bio_id"])

        bio_terms = [bio_by_id[bid].get("term", bid) for bid in bio_ids if bid in bio_by_id]

        if len(bio_terms) >= 2 or len(cand_list) >= 2:
            cross_bio_matches.append({
                "geometry_id": gid,
                "geometry_name": geo.get("geometry_name", gid),
                "section": geo.get("section", ""),
                "candidate_count": len(cand_list),
                "observation_count": len(obs_list),
                "bio_terms": bio_terms,
                "candidates": [c.get("name", c.get("candidate_id", "")) for c in cand_list],
                "match_features": geo.get("match_features", ""),
                "confidence": _avg_confidence(obs_list),
            })

    # Rank by frequency and confidence
    cross_bio_matches.sort(key=lambda x: (x["candidate_count"] + x["observation_count"], x["confidence"]), reverse=True)

    return {
        "generated_at": str(Path().stat().st_mtime),
        "geometry_count": len(geometries),
        "candidate_count": len(candidates),
        "observation_count": len(observations),
        "high_frequency_geometries": cross_bio_matches,
    }


def _avg_confidence(observations: list[dict]) -> float:
    if not observations:
        return 0.0
    scores = {"high": 1.0, "medium": 0.6, "low": 0.3}
    total = sum(scores.get(o.get("confidence", "").lower(), 0.0) for o in observations)
    return round(total / len(observations), 2)


def render_report(analysis: dict) -> str:
    lines = [
        "# Geometry Pattern Report",
        "",
        "Type: `agent-report`",
        "Agent: `geometry-pattern-agent`",
        "",
        "## Summary",
        "",
        f"- Total geometries: {analysis['geometry_count']}",
        f"- Total candidates: {analysis['candidate_count']}",
        f"- Total observations: {analysis['observation_count']}",
        "",
        "## High-Frequency Geometry Matches",
        "",
        "These geometries appear across multiple candidates or biological contexts. They are the primary pattern matches to investigate.",
        "",
    ]

    for match in analysis["high_frequency_geometries"]:
        lines.append(f"### {match['geometry_name']} (`{match['geometry_id']}`)")
        lines.append("")
        lines.append(f"- Section: {match['section']}")
        lines.append(f"- Candidates: {match['candidate_count']}")
        lines.append(f"- Observations: {match['observation_count']}")
        lines.append(f"- Average confidence: {match['confidence']}")
        lines.append(f"- Match features: {match['match_features']}")
        lines.append("")
        lines.append("**Candidates:**")
        for c in match["candidates"]:
            lines.append(f"- [[{c}]]")
        lines.append("")
        if match["bio_terms"]:
            lines.append("**Biological contexts:**")
            for b in match["bio_terms"]:
                lines.append(f"- [[{b}]]")
            lines.append("")
        lines.append("---")
        lines.append("")

    if not analysis["high_frequency_geometries"]:
        lines.append("No multi-candidate or multi-biology geometry matches found yet.")
        lines.append("")

    lines.append("## Method")
    lines.append("")
    lines.append("This report ignores biological categories until after geometric matches are identified. Biology is used only as confirmation metadata.")
    lines.append("")

    return "\n".join(lines)


def main() -> None:
    records = load_records(WORKBOOK)
    analysis = analyze_patterns(records)

    REPORT_DIR.mkdir(parents=True, exist_ok=True)

    # Write JSON for machine consumption
    json_path = REPORT_DIR / "geometry-pattern-report.json"
    json_path.write_text(json.dumps(analysis, indent=2), encoding="utf-8")

    # Write markdown for Obsidian
    md_path = REPORT_DIR / "geometry-pattern-report.md"
    md_path.write_text(render_report(analysis), encoding="utf-8")

    print(f"Report written to {md_path}")


if __name__ == "__main__":
    main()

